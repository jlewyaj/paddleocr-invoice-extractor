from openpyxl import Workbook
from pathlib import Path
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """Writes a list of Invoice objects out to an .xlsx spreadsheet."""

    def write(self, invoices, filename="Invoice_Extract.xlsx"):
        """Create a single-sheet workbook with one row per invoice.

        Empty invoices (all fields falsy, e.g. a page where extraction
        failed completely) are skipped. Column widths are auto-sized
        to fit their content. Output is saved under an "output" folder.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Invoices"

        # Header Row
        sheet.append([
            "Page",
            "Receipt No",
            "Hospital",
            "Doctor",
            "PRC License",
            "Patient",
            "Date",
            "Total",
        ])

        # Bold the header row
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        # Data Rows
        for invoice in invoices:
            # Skip invoices where every field is empty/zero/falsy —
            # these represent pages that yielded no useful data.
            if not any(vars(invoice).values()):
                continue

            sheet.append([
                invoice.page,
                invoice.receipt_no,
                invoice.hospital,
                invoice.doctor,
                invoice.prc_license,
                invoice.patient,
                invoice.date,
                invoice.total,
            ])

        # Auto-size each column's width based on its longest cell value
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            sheet.column_dimensions[column_letter].width = max_length + 2

        # Ensure the output directory exists, then save the workbook there
        output = Path("output")
        output.mkdir(exist_ok=True)
        
        workbook.save(output/filename)
